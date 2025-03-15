import re
import json
from pathlib import Path
from io import BytesIO
from collections import Counter
from flask import Blueprint, render_template, send_file
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font


stats_bp = Blueprint('statistics', __name__, template_folder='templates')


def calculate_corrections_count(original, corrected):
    orig_words = original.split()
    corr_words = corrected.split()
    count = sum(1 for ow, cw in zip(orig_words, corr_words) if ow != cw)
    count += abs(len(orig_words) - len(corr_words))
    return count


def gather_statistics():
    word_counter = Counter()
    object_counter = Counter()
    corrections = []
    data_folder = Path('data')    
    for json_file in data_folder.glob('*.json'):
        try:
            with json_file.open('r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            print(f"Ошибка чтения {json_file}: {e}")
            continue

        translated_text = data.get("translated_text", "")
        words = re.findall(r'\w+', translated_text.lower())
        word_counter.update(words)

        objects = data.get("objects_translated", [])
        object_counter.update(objects)

        original_text = data.get("text", "")
        corrected_text = data.get("corrected_text", "")
        corr_count = calculate_corrections_count(original_text, corrected_text)
        corrections.append((data.get("original_filename", json_file.name), corr_count))
    
    if word_counter:
        most_common_word, most_common_word_count = word_counter.most_common(1)[0]
    else:
        most_common_word, most_common_word_count = ("", 0)

    if object_counter:
        most_common_object, most_common_object_count = object_counter.most_common(1)[0]
    else:
        most_common_object, most_common_object_count = ("", 0)

    if corrections:
        max_corr_file, max_corr_count = max(corrections, key=lambda x: x[1])
    else:
        max_corr_file, max_corr_count = ("", 0)

    top_words = list(word_counter.most_common(10))
    top_objects = list(object_counter.most_common(10))

    return {
        "most_common_word": most_common_word,
        "most_common_word_count": most_common_word_count,
        "most_common_object": most_common_object,
        "most_common_object_count": most_common_object_count,
        "max_corr_file": max_corr_file,
        "max_corr_count": max_corr_count,
        "top_words": top_words,
        "top_objects": top_objects
    }


@stats_bp.route('/stats')
def stats_page():
    stats_data = gather_statistics()
    return render_template("stats.html", stats=stats_data)


def create_excel_stats():
    stats_data = gather_statistics()
    most_common_word = stats_data["most_common_word"]
    most_common_word_count = stats_data["most_common_word_count"]
    most_common_object = stats_data["most_common_object"]
    most_common_object_count = stats_data["most_common_object_count"]
    max_corr_file = stats_data["max_corr_file"]
    max_corr_count = stats_data["max_corr_count"]
    top_words = stats_data["top_words"]
    top_objects = stats_data["top_objects"]
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Общее"
    header_font = Font(bold=True)
    summary_sheet["A1"] = "Показатель"
    summary_sheet["B1"] = "Значение"
    summary_sheet["A1"].font = header_font
    summary_sheet["B1"].font = header_font
    summary_sheet.append(["Слово, встречающееся чаще всего", f"{most_common_word} ({most_common_word_count} раз)"])
    summary_sheet.append(["Объект, встречающийся чаще всего", f"{most_common_object} ({most_common_object_count} раз)"])
    summary_sheet.append(["Файл с наибольшим количеством исправлений", f"{max_corr_file} ({max_corr_count} исправлений)"])
    summary_sheet.column_dimensions["A"].width = 50
    summary_sheet.column_dimensions["B"].width = 30

    words_sheet = wb.create_sheet(title="Слова")
    words_sheet.append(["Слово", "Частота"])
    words_sheet["A1"].font = header_font
    words_sheet["B1"].font = header_font
    for word, count in top_words:
        words_sheet.append([word, count])
    words_sheet.column_dimensions["A"].width = 20
    words_sheet.column_dimensions["B"].width = 15

    words_chart = BarChart()
    words_chart.title = "Top 10 слов"
    words_chart.x_axis.title = "Слово"
    words_chart.y_axis.title = "Частота"
    data_ref = Reference(words_sheet, min_col=2, min_row=1, max_row=len(top_words) + 1)
    cats_ref = Reference(words_sheet, min_col=1, min_row=2, max_row=len(top_words) + 1)
    words_chart.add_data(data_ref, titles_from_data=True)
    words_chart.set_categories(cats_ref)
    words_chart.height = 10
    words_chart.width = 20
    words_sheet.add_chart(words_chart, "D2")

    objects_sheet = wb.create_sheet(title="Объект")
    objects_sheet.append(["Объект", "Частота"])
    objects_sheet["A1"].font = header_font
    objects_sheet["B1"].font = header_font
    for obj, count in top_objects:
        objects_sheet.append([obj, count])
    objects_sheet.column_dimensions["A"].width = 20
    objects_sheet.column_dimensions["B"].width = 15

    objects_chart = BarChart()
    objects_chart.title = "Top 10 объектов"
    objects_chart.x_axis.title = "Объект"
    objects_chart.y_axis.title = "Частота"
    data_ref2 = Reference(objects_sheet, min_col=2, min_row=1, max_row=len(top_objects) + 1)
    cats_ref2 = Reference(objects_sheet, min_col=1, min_row=2, max_row=len(top_objects) + 1)
    objects_chart.add_data(data_ref2, titles_from_data=True)
    objects_chart.set_categories(cats_ref2)
    objects_chart.height = 10
    objects_chart.width = 20
    objects_sheet.add_chart(objects_chart, "D2")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@stats_bp.route('/export_stats_excel')
def export_stats_excel():
    excel_stats = create_excel_stats()
    return send_file(
        excel_stats,
        as_attachment=True,
        download_name="statistics.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
